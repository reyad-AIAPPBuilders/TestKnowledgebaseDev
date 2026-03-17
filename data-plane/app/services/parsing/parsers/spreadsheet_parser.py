"""Spreadsheet parser for XLSX/XLS files using openpyxl."""

from openpyxl import load_workbook

from app.services.parsing.models import DocumentMetadata, DocumentType, ParseOptions, TableData
from app.services.parsing.parsers.base import BaseParser, ParsedContent
from app.utils.logger import get_logger

log = get_logger(__name__)


class SpreadsheetParser(BaseParser):
    """Extract text and tables from Excel files."""

    def supports(self) -> list[str]:
        return [DocumentType.XLSX.value, DocumentType.XLS.value]

    async def parse(self, file_path: str, options: ParseOptions) -> ParsedContent:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return self._extract(wb, options)
        finally:
            wb.close()

    def _extract(self, wb, options: ParseOptions) -> ParsedContent:
        text_parts: list[str] = []
        tables: list[TableData] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data: list[list[str]] = []
            text_parts.append(f"## Sheet: {sheet_name}")

            for row in ws.iter_rows(values_only=True):
                cells = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in cells):
                    rows_data.append(cells)
                    text_parts.append(" | ".join(cells))

            if options.extract_tables and len(rows_data) >= 1:
                tables.append(TableData(
                    headers=rows_data[0] if rows_data else [],
                    rows=rows_data[1:] if len(rows_data) > 1 else [],
                ))

        full_text = "\n".join(text_parts)

        props = wb.properties
        metadata = DocumentMetadata(
            title=props.title or None if props else None,
            author=props.creator or None if props else None,
            subject=props.subject or None if props else None,
            creation_date=str(props.created) if props and props.created else None,
            modification_date=str(props.modified) if props and props.modified else None,
            word_count=len(full_text.split()) if full_text else 0,
            char_count=len(full_text),
        )

        return ParsedContent(
            text=full_text,
            metadata=metadata,
            tables=tables,
            pages_parsed=len(wb.sheetnames),
            pages_failed=0,
        )
